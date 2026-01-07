import pandas as pd
from io import BytesIO
import streamlit as st
import file_merger
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

def save_and_download(merged, output_name):
    # Fix lỗi NaN
    merged = merged.fillna("")
    path = file_merger.save_file(merged, output_name)
    st.success(f"✅ Đã tạo file mới: {path}")

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False)
    buffer.seek(0)

    st.download_button(
        label="📥 Tải file kết quả",
        data=buffer.getvalue(),
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def save_with_form(merged, form_file, output_name, sheet_name=None, start_row=10, start_col=1):
    wb = load_workbook(form_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    merged = merged.fillna("")

    for r_idx, row in enumerate(merged.values.tolist(), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(cell, MergedCell):
                # ghi vào ô gốc của vùng merge
                for merged_range in ws.merged_cells.ranges:
                    if (r_idx, c_idx) in merged_range.cells:
                        top_left = merged_range.min_row, merged_range.min_col
                        ws.cell(row=top_left[0], column=top_left[1], value=value)
                        break
            else:
                cell.value = value

    wb.save(output_name)
    st.success(f"✅ Đã tạo file mới dựa trên form: {output_name}")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="📥 Tải file kết quả (giữ format form)",
        data=buffer.getvalue(),
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )